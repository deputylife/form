import logging
import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from aiogram.utils import executor

# 🔥 Токен Telegram-бота
TOKEN = "TOKEN_BOT"
# Путь к файлу с вопросами
EXCEL_FILE = "questions.xlsx"
# Путь к файлу для сохранения ответов
RESPONSES_FILE = "responses.xlsx"

# Логирование (чтобы видеть ошибки)
logging.basicConfig(level=logging.INFO)

# Создаем бота
bot = Bot(token=TOKEN)
dp = Dispatcher(bot)

# Храним вопросы и их варианты
questions = []
options = {}

# Храним ответы пользователей
user_answers = {}

# Клавиатура для начала опроса
start_keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
start_keyboard.add(KeyboardButton("Начать опрос"))

# 🟢 Команда /start
@dp.message_handler(commands=["start"])
async def start_survey(message: types.Message):
    user_id = message.from_user.id
    user_answers[user_id] = []  # Очищаем старые ответы
    await message.answer("Здравствуйте! Здесь вы можете протестировать как работает telegram-bot для опроса. Нажмите 'Начать опрос'.", reply_markup=start_keyboard)

# 🔄 Начинаем опрос
@dp.message_handler(lambda message: message.text == "Начать опрос")
async def ask_question(message: types.Message):
    user_id = message.from_user.id
    user_answers[user_id] = []  # Очищаем старые ответы перед началом
    
    # Загружаем вопросы из Excel
    load_questions_from_excel()
    
    await send_question(user_id, message)

# 📌 Функция отправки вопроса
async def send_question(user_id, message):
    current_question = len(user_answers[user_id])

    if current_question < len(questions):
        question_text = questions[current_question]
        question_type = options.get(question_text, {}).get("type", "text")

        # Если есть варианты ответа → добавляем кнопки
        if question_type == "Выбор" or question_type == "Множественный выбор":
            keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
            for option in options[question_text]["choices"]:
                keyboard.add(KeyboardButton(option))
            await message.answer(question_text, reply_markup=keyboard)
        else:
            # Если нет вариантов с кнопками — сбрасываем клавиатуру
            await message.answer(question_text, reply_markup=ReplyKeyboardRemove())
    else:
        await message.answer("Спасибо!")
        save_to_excel(user_id)
        await message.answer("✅ Данные сохранены и направлены в Минцифры России!", reply_markup=start_keyboard)
        await message.answer("Контакт для связи: @deputylife")

# 📝 Обработчик ответов
@dp.message_handler()
async def handle_answer(message: types.Message):
    user_id = message.from_user.id
    current_question = len(user_answers[user_id])

    if current_question < len(questions):
        user_answers[user_id].append(message.text)
        await send_question(user_id, message)

# 📡 Функция загрузки вопросов из Excel
def load_questions_from_excel():
    global questions, options
    questions = []
    options = {}

    # Открываем Excel файл с вопросами
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        question, question_type, option_str = row
        questions.append(question)
        
        if question_type == "Выбор" or question_type == "Множественный выбор":
            options[question] = {
                "type": question_type,
                "choices": option_str.split(', ')  # Разделяем варианты через запятую
            }
        else:
            options[question] = {"type": question_type}

# 📡 Функция записи ответов в новый Excel файл (responses.xlsx)
def save_to_excel(user_id):
    # Открываем или создаем файл для хранения ответов
    try:
        wb = openpyxl.load_workbook(RESPONSES_FILE)
        sheet = wb.active
    except FileNotFoundError:
        # Если файл не существует, создаем новый
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Добавляем заголовки (имена колонок)
        sheet.append(questions)  # Заголовки — вопросы

    # Преобразуем данные для записи в Excel
    data = user_answers[user_id]
    
    # Находим следующую пустую строку
    next_row = sheet.max_row + 1
    
    # Записываем каждый ответ в свою ячейку
    sheet.append(data)  # Ответы записываем по одному в каждую ячейку

    # Сохраняем файл
    wb.save(RESPONSES_FILE)

# 🚀 Запуск бота
if __name__ == "__main__":
    executor.start_polling(dp, skip_updates=True)
